"""
UiPath Process Health ETL Module

This module builds a unified process observability dataset by:
- extracting process inventory from UiPath Orchestrator
- enriching with business mapping rules
- analyzing job execution history
- parsing trigger schedules
- calculating process health scores and states

Outputs are used for:
- Power BI dashboards
- alert monitoring systems
- operational observability insights
"""

import os
import re
import sqlite3
import logging
from datetime import datetime
from openpyxl import load_workbook
import pyodbc
import pandas as pd
import pytz


# =========================================================
# GLOBALS
# =========================================================
# Local timezone used for all time-based calculations
LOCAL_TIMEZONE = pytz.timezone("Asia/Amman")

SUCCESS_STATES = ["successful"]
FAILURE_STATES = ["faulted", "failed", "stopped", "aborted"]


# =========================================================
# HELPERS
# =========================================================
def load_settings(configPath, SheetName, Key, Value):
    df = pd.read_excel(configPath, sheet_name=SheetName)
    df[Key] = df[Key].astype(str).str.strip()
    df[Value] = df[Value].astype(str).str.strip()
    return df.set_index(Key)[Value].to_dict()


def create_formatter(_format_string):
    return logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def create_log_file_handler(logFilePath, formatter):
    os.makedirs(os.path.dirname(logFilePath), exist_ok=True)
    fh = logging.FileHandler(logFilePath, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(formatter)
    return fh


def setup_loggers(log_file_path):
    formatter = create_formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fileHandler = create_log_file_handler(log_file_path, formatter)

    logging.getLogger('___main___').handlers.clear()
    logging.getLogger('___Init___').handlers.clear()
    logging.getLogger('___Processing___').handlers.clear()

    logger_main = logging.getLogger('___main___')
    logger_init = logging.getLogger('___Init___')
    logger_process = logging.getLogger('___Processing___')

    logger_main.setLevel(logging.DEBUG)
    logger_init.setLevel(logging.DEBUG)
    logger_process.setLevel(logging.DEBUG)

    if not logger_main.handlers:
        logger_main.addHandler(fileHandler)
    if not logger_init.handlers:
        logger_init.addHandler(fileHandler)
    if not logger_process.handlers:
        logger_process.addHandler(fileHandler)

    logger_main.propagate = False
    logger_init.propagate = False
    logger_process.propagate = False

    return logger_main, logger_init, logger_process, fileHandler


def build_sql_conn_str(server, database, username, password):
    return (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={server};DATABASE={database};UID={username};PWD={password}"
    )


def normalize_text(val):
    if pd.isna(val):
        return ""
    return str(val).strip()


def normalize_yes_no(val):
    if pd.isna(val):
        return ""
    v = str(val).strip().lower()
    if v in ["yes", "y", "true", "1", "1.0"]:
        return "Yes"
    if v in ["no", "n", "false", "0", "0.0"]:
        return "No"
    return ""


def normalize_process_name(val):
    if pd.isna(val):
        return ""
    v = str(val).strip().lower()
    v = re.sub(r'[^a-z0-9]+', '', v)
    return v


def classify_usage(days_since_last_activity):
    if pd.isna(days_since_last_activity):
        return "No Activity History"
    elif days_since_last_activity >= 90:
        return "Retirement Candidate"
    elif days_since_last_activity >= 60:
        return "Stopped"
    elif days_since_last_activity >= 30:
        return "Low Usage"
    else:
        return "Active"



def save_uipath_sheet_only(workbook_path, df_uipath, logger, sheet_name="UIPATH"):
    """
    Update only the UIPATH sheet inside an existing workbook.
    Keep all other sheets unchanged.
    If UIPATH sheet exists, replace it.
    If workbook does not exist, create it.
    """
    os.makedirs(os.path.dirname(workbook_path), exist_ok=True)

    # clean columns
    export_df = df_uipath.copy()
    export_df.columns = export_df.columns.str.strip()

    if os.path.exists(workbook_path):
        # load existing workbook and preserve other sheets
        book = load_workbook(workbook_path)

        if sheet_name in book.sheetnames:
            std = book[sheet_name]
            book.remove(std)

        with pd.ExcelWriter(
            workbook_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

        logger.info(f"{sheet_name} sheet updated in existing workbook: {workbook_path}")

    else:
        # create new workbook if file does not exist
        with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="w") as writer:
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

        logger.info(f"Workbook created and {sheet_name} sheet written: {workbook_path}")
# =========================================================
# BUSINESS INPUT
# =========================================================
def load_business_mapping(path, logger):
    df = pd.read_excel(path, sheet_name="UIPATH")
    df.columns = [str(c).strip() for c in df.columns]

    rename_map = {
        "process_name": "ProcessName",
        "severity": "Severity",
        "critical_once_daily": "critical_once_daily",
        "ticket_threshold_count": "ticket_threshold_count",
        "StaleJobsIgnore": "StaleJobsIgnore"
    }
    df = df.rename(columns=rename_map)

    if "ProcessName" not in df.columns:
        raise ValueError("Business file must contain ProcessName column")

    df["ProcessName"] = df["ProcessName"].fillna("").astype(str).str.strip()
    df["ProcessName"] = df["ProcessName"].replace(["nan", "None", "<NA>"], "")
    df = df[df["ProcessName"] != ""].copy()

    df["ProcessNameKey"] = df["ProcessName"].apply(normalize_process_name)

    if "Severity" not in df.columns:
        df["Severity"] = "UNKNOWN"
    df["Severity"] = df["Severity"].fillna("UNKNOWN").astype(str).str.strip().str.upper()

    if "critical_once_daily" not in df.columns:
        df["critical_once_daily"] = "NO"
    df["critical_once_daily"] = df["critical_once_daily"].fillna("NO").astype(str).str.strip().str.upper()

    if "ticket_threshold_count" not in df.columns:
        df["ticket_threshold_count"] = 3
    df["ticket_threshold_count"] = pd.to_numeric(
        df["ticket_threshold_count"], errors="coerce"
    ).fillna(3).astype(int)

    if "StaleJobsIgnore" not in df.columns:
        df["StaleJobsIgnore"] = "NO"
    df["StaleJobsIgnore"] = df["StaleJobsIgnore"].fillna("NO").astype(str).str.strip().str.upper()

    df = df.drop_duplicates(subset=["ProcessNameKey"]).copy()

    logger.info(f"Business mapping loaded | rows={len(df)}")
    return df

# =========================================================
# JOBS FROM CURATED
# =========================================================
def load_jobs_from_curated(curated_db, logger, lookback_days=None):
    query = """
    SELECT
        releaseId as ReleaseId,
        processName as ProcessName,
        jobStart,
        jobEnd,
        jobState as State,
        robotName as RobotName,
        machineName as MachineName
    FROM jobs_all
    """
    if lookback_days is not None:
        query += f"""
        WHERE datetime(jobStart) >= datetime('now', '-{int(lookback_days)} day')
        """

    conn = sqlite3.connect(curated_db)
    df = pd.read_sql(query, conn)
    conn.close()

    if df.empty:
        logger.warning("No rows returned from curated jobs_all")
        return pd.DataFrame(columns=[
            "ReleaseId", "ProcessName", "ProcessNameKey", "jobStart", "jobEnd",
            "State", "RobotName", "MachineName", "RunTime", "DurationMinutes"
        ])

    df["ProcessName"] = df["ProcessName"].astype(str).str.strip()
    df["ProcessNameKey"] = df["ProcessName"].apply(normalize_process_name)
    df["State"] = df["State"].astype(str).str.strip().str.lower()
    df["RobotName"] = df["RobotName"].astype(str).str.strip()
    df["MachineName"] = df["MachineName"].astype(str).str.strip()
    df["ReleaseId"] = pd.to_numeric(df["ReleaseId"], errors="coerce")

    df["jobStart"] = pd.to_datetime(df["jobStart"], errors="coerce")
    df["jobEnd"] = pd.to_datetime(df["jobEnd"], errors="coerce")
    df["RunTime"] = df["jobEnd"].fillna(df["jobStart"])

    df["DurationMinutes"] = (df["jobEnd"] - df["jobStart"]).dt.total_seconds() / 60
    df.loc[df["DurationMinutes"] < 0, "DurationMinutes"] = pd.NA
    df["DurationMinutes"] = pd.to_numeric(df["DurationMinutes"], errors="coerce")

    df["State"] = df["State"].astype("category")
    df["ProcessName"] = df["ProcessName"].astype("string")
    df["RobotName"] = df["RobotName"].astype("string")
    df["MachineName"] = df["MachineName"].astype("string")

    logger.info(f"Jobs loaded from curated lake | rows={len(df)}")
    return df


# =========================================================
# RELEASES / QUEUES / SCHEDULES
# =========================================================
def extract_releases(sql_conn, logger):
    sql_query = """
    SELECT
        Id,
        Name,
        ProcessKey,
        OrganizationUnitId,
        IsDeleted
    FROM Releases
    WHERE ISNULL(IsDeleted, 0) = 0
    """
    df = pd.read_sql(sql_query, sql_conn)

    df = df.rename(columns={
        "Id": "ReleaseId",
        "Name": "ReleaseName"
    })

    df["ReleaseId"] = pd.to_numeric(df["ReleaseId"], errors="coerce")
    df["ReleaseName"] = df["ReleaseName"].astype(str).str.strip()
    df["ProcessNameKey"] = df["ReleaseName"].apply(normalize_process_name)

    logger.info(f"Releases extracted | rows={len(df)}")
    return df


def extract_queue_summary(sql_conn, logger):
    sql_query = """
    SELECT
        qd.Id AS QueueDefinitionId,
        qd.Name AS QueueName,
        qd.ReleaseId,
        qd.OrganizationUnitId,
        MAX(qi.LastModificationTime) AS LastQueueTransactionTime,
        COUNT(*) AS TotalItems,
        SUM(CASE WHEN qi.Status = 0 THEN 1 ELSE 0 END) AS NewItems,
        SUM(CASE WHEN qi.Status = 1 THEN 1 ELSE 0 END) AS InProgressItems,
        SUM(CASE WHEN qi.Status = 2 THEN 1 ELSE 0 END) AS FailedQueueItems,
        SUM(CASE WHEN qi.Status = 3 THEN 1 ELSE 0 END) AS SuccessfulQueueItems,
        SUM(CASE WHEN qi.Status = 4 THEN 1 ELSE 0 END) AS AbandonedItems,
        SUM(CASE WHEN qi.Status = 5 THEN 1 ELSE 0 END) AS RetriedItems,
        SUM(CASE WHEN qi.Status IN (0,1) THEN 1 ELSE 0 END) AS PendingItems,
        GETDATE() AS QueryTime
    FROM QueueItems qi
    INNER JOIN QueueDefinitions qd
        ON qi.QueueDefinitionId = qd.Id
    WHERE ISNULL(qi.IsDeleted, 0) = 0
      AND ISNULL(qd.IsDeleted, 0) = 0
    GROUP BY qd.Id, qd.Name, qd.ReleaseId, qd.OrganizationUnitId
    """
    df = pd.read_sql(sql_query, sql_conn)

    df["QueueDefinitionId"] = pd.to_numeric(df["QueueDefinitionId"], errors="coerce")
    df["ReleaseId"] = pd.to_numeric(df["ReleaseId"], errors="coerce")
    df["QueueName"] = df["QueueName"].astype(str).str.strip()
    df["LastQueueTransactionTime"] = pd.to_datetime(df["LastQueueTransactionTime"], errors="coerce")
    df["QueryTime"] = pd.to_datetime(df["QueryTime"], errors="coerce")

    numeric_cols = [
        "TotalItems", "NewItems", "InProgressItems", "FailedQueueItems",
        "SuccessfulQueueItems", "AbandonedItems", "RetriedItems", "PendingItems"
    ]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    logger.info(f"Queue summary extracted | rows={len(df)}")
    return df


def extract_process_schedule_map(sql_conn, logger):
    sql_query = """
    SELECT
        ps.Id AS ProcessScheduleId,
        ps.ReleaseId,
        ps.Enabled AS TriggerEnabled,
        ps.StartProcessCron,
        ps.StartProcessCronDetails,
        ps.LastSuccessfulTime,
        ps.LastFailureTime,
        ps.TotalSuccessful,
        ps.TotalFailures,
        ps.QueueDefinitionId,
        ps.ItemsActivationThreshold,
        ps.ItemsPerJobActivationTarget,
        ps.MaxJobsForActivation,
        qd.Name AS QueueName
    FROM ProcessSchedules ps
    LEFT JOIN QueueDefinitions qd
        ON ps.QueueDefinitionId = qd.Id
    WHERE ISNULL(ps.IsDeleted, 0) = 0
    """
    df = pd.read_sql(sql_query, sql_conn)

    df["ProcessScheduleId"] = pd.to_numeric(df["ProcessScheduleId"], errors="coerce")
    df["ReleaseId"] = pd.to_numeric(df["ReleaseId"], errors="coerce")
    df["QueueDefinitionId"] = pd.to_numeric(df["QueueDefinitionId"], errors="coerce")
    df["TriggerEnabled"] = pd.to_numeric(df["TriggerEnabled"], errors="coerce").fillna(0)

    for col in ["LastSuccessfulTime", "LastFailureTime"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    for col in ["StartProcessCron", "StartProcessCronDetails", "QueueName"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).str.strip()

    # create empty summary column so downstream code does not break
    if "StartProcessCronSummary" not in df.columns:
        df["StartProcessCronSummary"] = ""

    df["TriggerType"] = df["QueueDefinitionId"].apply(lambda x: "Queue" if pd.notna(x) else "Time")

    logger.info(f"ProcessSchedules extracted | rows={len(df)}")
    return df

# =========================================================
# TRIGGER INTERVAL PARSER
# =========================================================
def parse_interval_minutes(summary_text, cron_text=""):
    s = str(summary_text or "").strip().lower()
    c = str(cron_text or "").strip()

    ordered_days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    def parse_num_list(text):
        vals = []
        for x in str(text).split(","):
            x = x.strip()
            if re.fullmatch(r"\d+", x):
                vals.append(int(x))
            else:
                return None
        return sorted(set(vals))

    def parse_hour_field(hour_text):
        hour_text = str(hour_text).strip()

        if hour_text == "*":
            return list(range(24)), "all_hours"

        m = re.fullmatch(r"(?:\*|\d+)/(\d+)", hour_text)
        if m:
            step = int(m.group(1))
            if step > 0:
                return list(range(0, 24, step)), f"every_{step}_hours"

        m = re.fullmatch(r"(\d{1,2})-(\d{1,2})", hour_text)
        if m:
            start_h = int(m.group(1))
            end_h = int(m.group(2))
            if 0 <= start_h <= 23 and 0 <= end_h <= 23:
                if start_h <= end_h:
                    return list(range(start_h, end_h + 1)), "hour_range"
                else:
                    return list(range(start_h, 24)) + list(range(0, end_h + 1)), "hour_range_wrap"

        if "," in hour_text:
            vals = parse_num_list(hour_text)
            if vals is not None:
                return vals, "hour_list"

        if re.fullmatch(r"\d{1,2}", hour_text):
            h = int(hour_text)
            if 0 <= h <= 23:
                return [h], "single_hour"

        return None, None

    def parse_minute_field(minute_text):
        minute_text = str(minute_text).strip()

        if minute_text == "*":
            return list(range(60)), 1, 60, "every_minute"

        m = re.fullmatch(r"(?:\*|\d+)/(\d+)", minute_text)
        if m:
            step = int(m.group(1))
            if step > 0:
                mins = list(range(0, 60, step))
                return mins, step, len(mins), f"every_{step}_minutes"

        m = re.fullmatch(r"(\d{1,2})-(\d{1,2})", minute_text)
        if m:
            start_m = int(m.group(1))
            end_m = int(m.group(2))
            if 0 <= start_m <= 59 and 0 <= end_m <= 59 and start_m <= end_m:
                mins = list(range(start_m, end_m + 1))
                return mins, 1, len(mins), "minute_range"

        if "," in minute_text:
            mins = parse_num_list(minute_text)
            if mins:
                if len(mins) > 1:
                    diffs = [mins[i + 1] - mins[i] for i in range(len(mins) - 1)]
                    wrap_diff = (60 - mins[-1]) + mins[0]
                    diffs.append(wrap_diff)
                    interval = diffs[0] if len(set(diffs)) == 1 else min(diffs)
                else:
                    interval = 60
                return mins, interval, len(mins), "minute_list"

        if re.fullmatch(r"\d{1,2}", minute_text):
            mval = int(minute_text)
            if 0 <= mval <= 59:
                return [mval], 60, 1, "fixed_minute"

        return None, None, None, None

    def parse_day_of_week_field(dow_text):
        dow_text = str(dow_text or "").strip().upper()

        if dow_text in ("*", "?"):
            return ordered_days.copy(), "all_days"

        m = re.fullmatch(r"(SUN|MON|TUE|WED|THU|FRI|SAT)-(SUN|MON|TUE|WED|THU|FRI|SAT)", dow_text)
        if m:
            start_d, end_d = m.group(1), m.group(2)
            i1, i2 = ordered_days.index(start_d), ordered_days.index(end_d)
            if i1 <= i2:
                return ordered_days[i1:i2 + 1], "day_range"
            else:
                return ordered_days[i1:] + ordered_days[:i2 + 1], "day_range_wrap"

        if "," in dow_text:
            days = []
            for part in dow_text.split(","):
                part = part.strip().upper()
                if part in ordered_days:
                    days.append(part)
                else:
                    return None, None
            return sorted(set(days), key=lambda d: ordered_days.index(d)), "day_list"

        if dow_text in ordered_days:
            return [dow_text], "single_day"

        return None, None

    def build_result(interval_minutes, runs_per_day, expected_interval_hours, schedule_type, active_hours, active_days):
        return (
            interval_minutes,
            runs_per_day,
            expected_interval_hours,
            schedule_type,
            active_hours,
            active_days
        )

    m = re.search(r"every\s+(\d+)\s+minutes?", s)
    if m:
        interval = int(m.group(1))
        return build_result(interval, round(1440 / interval, 4), round(interval / 60, 4), f"every_{interval}_minutes", "00-23", "SUN,MON,TUE,WED,THU,FRI,SAT")

    m = re.search(r"every\s+(\d+)\s+hours?", s)
    if m:
        interval = int(m.group(1)) * 60
        return build_result(interval, round(1440 / interval, 4), round(interval / 60, 4), f"every_{int(m.group(1))}_hours", "00-23", "SUN,MON,TUE,WED,THU,FRI,SAT")

    if "every minute" in s:
        return build_result(1, 1440, 0.0167, "every_minute", "00-23", "SUN,MON,TUE,WED,THU,FRI,SAT")

    if "hourly" in s or "every hour" in s:
        return build_result(60, 24, 1, "hourly", "00-23", "SUN,MON,TUE,WED,THU,FRI,SAT")

    if "daily" in s or "every day" in s:
        return build_result(1440, 1, 24, "daily", "00-23", "SUN,MON,TUE,WED,THU,FRI,SAT")

    parts = c.split()
    if len(parts) not in (6, 7):
        return build_result(None, None, None, "unknown", None, None)

    if len(parts) == 6:
        sec, minute, hour, day_month, month, day_week = parts
    else:
        sec, minute, hour, day_month, month, day_week, year = parts

    minute_values, minute_interval, minute_runs_per_hour, minute_type = parse_minute_field(minute)
    hour_values, hour_type = parse_hour_field(hour)
    active_day_values, day_type = parse_day_of_week_field(day_week)

    if minute_values is None or hour_values is None:
        return build_result(None, None, None, "unknown", None, None)

    runs_per_active_day = minute_runs_per_hour * len(hour_values)

    if active_day_values is None:
        active_days_text = None
        expected_runs_per_day = runs_per_active_day
    else:
        active_days_text = ",".join(active_day_values)
        if len(active_day_values) == 7:
            expected_runs_per_day = runs_per_active_day
        else:
            expected_runs_per_day = round((runs_per_active_day * len(active_day_values)) / 7, 4)

    is_specific_day_of_month = day_month not in ("*", "?", "1/1")

    if is_specific_day_of_month:
        schedule_type = "monthly"
    elif minute_type == "every_minute" and hour_type == "all_hours":
        schedule_type = "every_minute"
    elif minute_type and minute_type.startswith("every_") and hour_type == "all_hours":
        schedule_type = minute_type
    elif minute_type == "fixed_minute" and hour_type == "all_hours":
        schedule_type = "hourly"
    elif minute_type == "fixed_minute" and hour_type in ("hour_range", "hour_range_wrap", "hour_list"):
        schedule_type = "hourly_selected_hours"
    elif minute_type == "fixed_minute" and hour_type == "single_hour":
        schedule_type = "daily"
    else:
        schedule_type = "custom_cron"

    if active_day_values is not None and len(active_day_values) < 7 and schedule_type != "monthly":
        schedule_type = f"{schedule_type}_selected_days"

    active_hours = ",".join([f"{h:02d}" for h in hour_values])

    interval_minutes = None

    if schedule_type == "every_minute":
        interval_minutes = 1
    elif schedule_type.startswith("every_") and "minutes" in schedule_type:
        m = re.search(r"every_(\d+)_minutes", schedule_type)
        if m:
            interval_minutes = int(m.group(1))
    elif schedule_type.startswith("every_") and "hours" in schedule_type:
        m = re.search(r"every_(\d+)_hours", schedule_type)
        if m:
            interval_minutes = int(m.group(1)) * 60
    elif schedule_type == "hourly":
        interval_minutes = 60
    elif schedule_type == "hourly_selected_hours":
        if len(hour_values) > 1:
            diffs = [hour_values[i + 1] - hour_values[i] for i in range(len(hour_values) - 1)]
            wrap_diff = (24 - hour_values[-1]) + hour_values[0]
            diffs.append(wrap_diff)
            interval_minutes = min(diffs) * 60
        else:
            interval_minutes = 1440
    elif schedule_type == "daily":
        interval_minutes = 1440
    elif schedule_type.endswith("_selected_days"):
        if active_day_values and len(active_day_values) > 0:
            interval_minutes = int(round(10080 / len(active_day_values)))
        else:
            interval_minutes = 10080
    elif schedule_type == "monthly":
        interval_minutes = 43200
    elif schedule_type == "custom_cron":
        if expected_runs_per_day and expected_runs_per_day > 0:
            interval_minutes = round(1440 / expected_runs_per_day, 4)

    expected_interval_hours = round(interval_minutes / 60, 4) if interval_minutes is not None else None

    return build_result(
        interval_minutes,
        expected_runs_per_day,
        expected_interval_hours,
        schedule_type,
        active_hours,
        active_days_text
    )


def enrich_schedule_fields(df, logger):
    parsed = df.apply(
        lambda row: parse_interval_minutes(
            row.get("StartProcessCronSummary", ""),
            row.get("StartProcessCron", "")
        ),
        axis=1
    )

    parsed_df = pd.DataFrame(
        parsed.tolist(),
        columns=[
            "ExpectedIntervalMinutes",
            "ExpectedRunsPerDay",
            "ExpectedIntervalHours",
            "ScheduleType",
            "ActiveHours",
            "ActiveDays"
        ]
    )

    result = pd.concat([df.reset_index(drop=True), parsed_df], axis=1)
    logger.info(f"Schedule fields enriched | rows={len(result)}")
    return result


# =========================================================
# PROCESS MASTER MAPPING
# =========================================================
def build_process_trigger_map(df_schedules, df_releases, logger):
    df_map = df_schedules.merge(
        df_releases[["ReleaseId", "ReleaseName", "ProcessNameKey"]],
        on="ReleaseId",
        how="left"
    )

    df_map["ProcessName"] = df_map["ReleaseName"]
    df_map["MappingMethod"] = df_map["ProcessName"].notna().map({
        True: "ProcessSchedules.ReleaseId -> Releases.Id",
        False: "Unmapped"
    })

    logger.info(f"Process trigger map built | rows={len(df_map)}")
    return df_map


def build_master_process_mapping(df_releases, df_business, df_trigger_map, logger):
    # =====================================================
    # Start from ALL releases, not from business mapping
    # =====================================================
    df_master = df_releases.copy()

    df_master["ProcessName"] = df_master["ReleaseName"].fillna("").astype(str).str.strip()
    df_master["ProcessNameKey"] = df_master["ProcessName"].apply(normalize_process_name)

    # =====================================================
    # Merge business mapping
    # =====================================================
    business_cols = [
        "ProcessNameKey",
        "Severity",
        "critical_once_daily",
        "ticket_threshold_count",
        "StaleJobsIgnore"
    ]

    df_master = df_master.merge(
        df_business[business_cols].drop_duplicates(subset=["ProcessNameKey"]),
        on="ProcessNameKey",
        how="left"
    )

    # defaults for processes not found in business sheet
    df_master["Severity"] = df_master["Severity"].fillna("UNKNOWN").astype(str).str.strip().str.upper()
    df_master["critical_once_daily"] = df_master["critical_once_daily"].fillna("NO").astype(str).str.strip().str.upper()
    df_master["ticket_threshold_count"] = pd.to_numeric(
        df_master["ticket_threshold_count"], errors="coerce"
    ).fillna(3).astype(int)
    df_master["StaleJobsIgnore"] = df_master["StaleJobsIgnore"].fillna("NO").astype(str).str.strip().str.upper()

    df_master["InBusinessMapping"] = (df_master["Severity"] != "UNKNOWN").astype(int)
    df_master["MissingBusinessMapping"] = (df_master["Severity"] == "UNKNOWN").astype(int)

    # =====================================================
    # Aggregate latest trigger info by ProcessNameKey
    # =====================================================
    trigger_latest = df_trigger_map.copy()
    trigger_latest["HasQueueTrigger"] = trigger_latest["QueueDefinitionId"].notna().astype(int)

    trigger_agg = (
        trigger_latest
        .sort_values(["ProcessNameKey", "TriggerEnabled", "ProcessScheduleId"], ascending=[True, False, False])
        .groupby("ProcessNameKey", as_index=False)
        .agg(
            TriggerTypeLatest=("TriggerType", lambda x: "Queue" if "Queue" in list(x.dropna()) else ("Time" if len(list(x.dropna())) > 0 else "")),
            TriggerEnabledLatest=("TriggerEnabled", "max"),
            QueueName=("QueueName", lambda x: ", ".join(sorted(set([str(v).strip() for v in x.dropna() if str(v).strip()])))),
            QueueDefinitionId=("QueueDefinitionId", "max"),
            HasQueueTrigger=("HasQueueTrigger", "max"),
            ProcessScheduleId=("ProcessScheduleId", "max"),
            ItemsActivationThreshold=("ItemsActivationThreshold", "max"),
            ItemsPerJobActivationTarget=("ItemsPerJobActivationTarget", "max"),
            MaxJobsForActivation=("MaxJobsForActivation", "max"),
            configStartProcessCron=("StartProcessCron", "first"),
            configStartProcessCronSummary=("StartProcessCronSummary", "first"),
            configStartProcessCronDetails=("StartProcessCronDetails", "first"),
            ExpectedIntervalMinutes=("ExpectedIntervalMinutes", "first"),
            ExpectedRunsPerDay=("ExpectedRunsPerDay", "first"),
            ExpectedIntervalHours=("ExpectedIntervalHours", "first"),
            ScheduleType=("ScheduleType", "first"),
            ActiveHours=("ActiveHours", "first"),
            ActiveDays=("ActiveDays", "first")
        )
    )

    df_master = df_master.merge(trigger_agg, on="ProcessNameKey", how="left")

    df_master["Source"] = "UIPATH_PROCESS"

    logger.info(f"Full master process mapping built | rows={len(df_master)}")
    logger.info(f"Processes missing from business mapping | rows={int(df_master['MissingBusinessMapping'].sum())}")

    return df_master


# =========================================================
# JOB SUMMARY
# =========================================================
def build_jobs_summary(df_jobs, logger):
    df_jobs = df_jobs.copy()
    now_naive = pd.Timestamp.now(tz=LOCAL_TIMEZONE).tz_localize(None)

    df_jobs["ReleaseId"] = pd.to_numeric(df_jobs["ReleaseId"], errors="coerce")
    df_jobs = df_jobs[df_jobs["ReleaseId"].notna()].copy()
    df_jobs["ReleaseId"] = df_jobs["ReleaseId"].astype("int32")

    latest_name_base = df_jobs[df_jobs["jobStart"].notna()].copy()
    if latest_name_base.empty:
        latest_release_name = pd.DataFrame(columns=["ReleaseId", "ProcessName_job", "ProcessNameKey"])
    else:
        idx_latest_name = latest_name_base.groupby("ReleaseId")["jobStart"].idxmax()
        latest_release_name = (
            latest_name_base.loc[idx_latest_name, ["ReleaseId", "ProcessName", "ProcessNameKey"]]
            .drop_duplicates(subset=["ReleaseId"])
            .rename(columns={"ProcessName": "ProcessName_job"})
        )

    last_run = df_jobs.groupby("ReleaseId", as_index=False).agg(LastRunTime=("RunTime", "max"))

    success_df = df_jobs[df_jobs["State"].isin(SUCCESS_STATES)].copy()
    last_success = success_df.groupby("ReleaseId", as_index=False).agg(LastSuccessTime=("RunTime", "max"))

    failure_df = df_jobs[df_jobs["State"].isin(FAILURE_STATES)].copy()
    last_failure = failure_df.groupby("ReleaseId", as_index=False).agg(LastFailureTime=("RunTime", "max"))

    last_seen_base = df_jobs[df_jobs["RunTime"].notna()].copy()
    if last_seen_base.empty:
        last_seen = pd.DataFrame(columns=["ReleaseId", "RobotName", "MachineName", "LastJobState"])
    else:
        idx_last_seen = last_seen_base.groupby("ReleaseId")["RunTime"].idxmax()
        last_seen = (
            last_seen_base.loc[idx_last_seen, ["ReleaseId", "RobotName", "MachineName", "State"]]
            .drop_duplicates(subset=["ReleaseId"])
            .rename(columns={"State": "LastJobState"})
        )

    df_7d = df_jobs[df_jobs["RunTime"] >= (now_naive - pd.Timedelta(days=7))].copy()

    counts_7d = (
        df_7d.groupby("ReleaseId", as_index=False)
        .agg(
            RunCount7D=("ReleaseId", "count"),
            SuccessCount7D=("State", lambda x: x.isin(SUCCESS_STATES).sum()),
            FailureCount7D=("State", lambda x: x.isin(FAILURE_STATES).sum())
        )
    )

    df_7d["DurationMinutes"] = pd.to_numeric(df_7d["DurationMinutes"], errors="coerce")
    duration_base = df_7d[df_7d["DurationMinutes"].notna()].copy()

    if duration_base.empty:
        duration_stats = pd.DataFrame(columns=["ReleaseId", "AvgDurationMinutes7D", "MaxDurationMinutes7D"])
    else:
        duration_stats = (
            duration_base.groupby("ReleaseId", as_index=False)
            .agg(
                AvgDurationMinutes7D=("DurationMinutes", "mean"),
                MaxDurationMinutes7D=("DurationMinutes", "max")
            )
        )

    running_df = df_jobs[df_jobs["State"] == "running"].copy()
    running_flag = (
        running_df.groupby("ReleaseId", as_index=False)
        .agg(IsCurrentlyRunning=("ReleaseId", "count"))
    )
    if not running_flag.empty:
        running_flag["IsCurrentlyRunning"] = 1

    final = latest_release_name.merge(last_run, on="ReleaseId", how="left")
    final = final.merge(last_success, on="ReleaseId", how="left")
    final = final.merge(last_failure, on="ReleaseId", how="left")
    final = final.merge(last_seen, on="ReleaseId", how="left")
    final = final.merge(counts_7d, on="ReleaseId", how="left")
    final = final.merge(duration_stats, on="ReleaseId", how="left")
    final = final.merge(running_flag, on="ReleaseId", how="left")

    for col in ["RunCount7D", "SuccessCount7D", "FailureCount7D", "IsCurrentlyRunning"]:
        if col not in final.columns:
            final[col] = 0
        final[col] = pd.to_numeric(final[col], errors="coerce").fillna(0)

    for col in ["AvgDurationMinutes7D", "MaxDurationMinutes7D"]:
        if col not in final.columns:
            final[col] = pd.NA
        final[col] = pd.to_numeric(final[col], errors="coerce")

    logger.info(f"Jobs summary built | rows={len(final)}")
    return final


# =========================================================
# HEALTH LOGIC
# =========================================================
def classify_process_state(row):
    expected = pd.to_numeric(row.get("ExpectedIntervalMinutes"), errors="coerce")
    minutes_since_last_run = row.get("MinutesSinceLastRun")
    minutes_since_last_success = row.get("MinutesSinceLastSuccess")
    minutes_since_last_queue_txn = row.get("MinutesSinceLastQueueTransaction")
    days_since_last_activity = row.get("DaysSinceLastActivity")

    failure_count_7d = pd.to_numeric(row.get("FailureCount7D", 0), errors="coerce")
    success_count_7d = pd.to_numeric(row.get("SuccessCount7D", 0), errors="coerce")
    pending_items = pd.to_numeric(row.get("PendingItems", 0), errors="coerce")

    is_enabled = True
    trigger_enabled = pd.to_numeric(row.get("TriggerEnabledLatest", 1), errors="coerce")
    trigger_type = normalize_text(row.get("TriggerTypeLatest")).lower()
    has_job_history = pd.to_numeric(row.get("HasJobHistory", 0), errors="coerce")
    has_queue_trigger = pd.to_numeric(row.get("HasQueueTrigger", 0), errors="coerce")
    is_running = pd.to_numeric(row.get("IsCurrentlyRunning", 0), errors="coerce")

    has_queue_history = pd.notna(minutes_since_last_queue_txn)
    has_any_history = (has_job_history == 1) or has_queue_history
    is_queue_process = (has_queue_trigger == 1) or (trigger_type == "queue")

    if not is_enabled:
        return "Disabled"

    if is_queue_process:
        if is_running == 1:
            return "Running"

        if pd.notna(days_since_last_activity) and days_since_last_activity >= 60:
            return "Stopped" if has_any_history else "LongTime Not Used"

        if pd.notna(minutes_since_last_queue_txn):
            if pending_items <= 5:
                return "Healthy"
            elif pending_items <= 20:
                return "Warning" if minutes_since_last_queue_txn > 360 else "Healthy"
            elif pending_items <= 70:
                return "At Risk" if minutes_since_last_queue_txn > 720 else "Warning"
            else:
                return "At Risk" if minutes_since_last_queue_txn > 720 else "Warning"

        if pd.notna(trigger_enabled) and trigger_enabled == 0:
            return "Stopped" if has_any_history else "LongTime Not Used"

        return "No Job History"

    if pd.notna(trigger_enabled) and trigger_enabled == 0:
        return "Stopped" if has_any_history else "LongTime Not Used"

    if pd.notna(days_since_last_activity) and days_since_last_activity >= 60:
        return "Stopped" if has_any_history else "LongTime Not Used"

    if pd.isna(expected) or expected <= 0:
        if is_running == 1:
            return "Running"
        if has_job_history == 0:
            return "No Job History"
        return "Unknown"

    if is_running == 1:
        return "Running"

    if pd.isna(minutes_since_last_run):
        return "No Job History"

    avg_duration = pd.to_numeric(row.get("AvgDurationMinutes7D", 0), errors="coerce")
    max_duration = pd.to_numeric(row.get("MaxDurationMinutes7D", 0), errors="coerce")

    duration_buffer = max(avg_duration if pd.notna(avg_duration) else 0, 15)
    major_buffer = max(max_duration if pd.notna(max_duration) else duration_buffer, duration_buffer)

    warning_threshold = expected + duration_buffer + 100
    down_threshold = expected + major_buffer + 280

    if minutes_since_last_run > down_threshold:
        return "At Risk"
    if minutes_since_last_run > warning_threshold:
        return "Warning"

    if pd.notna(minutes_since_last_success):
        if minutes_since_last_success > down_threshold:
            return "At Risk"
        if minutes_since_last_success > warning_threshold:
            return "Warning"

    if failure_count_7d >= 10 and success_count_7d == 0:
        return "At Risk"
    if failure_count_7d >= 7:
        return "Warning"

    return "Healthy"


def calculate_health_score(row):
    score = 100

    expected = pd.to_numeric(row.get("ExpectedIntervalMinutes"), errors="coerce")
    minutes_since_last_run = row.get("MinutesSinceLastRun")
    minutes_since_last_success = row.get("MinutesSinceLastSuccess")
    minutes_since_last_queue_txn = row.get("MinutesSinceLastQueueTransaction")
    days_since_last_activity = row.get("DaysSinceLastActivity")
    pending_items = pd.to_numeric(row.get("PendingItems", 0), errors="coerce")
    failure_count_7d = pd.to_numeric(row.get("FailureCount7D", 0), errors="coerce")

    trigger_enabled = pd.to_numeric(row.get("TriggerEnabledLatest", 1), errors="coerce")
    trigger_type = normalize_text(row.get("TriggerTypeLatest")).lower()
    has_queue_trigger = pd.to_numeric(row.get("HasQueueTrigger", 0), errors="coerce")
    is_running = pd.to_numeric(row.get("IsCurrentlyRunning", 0), errors="coerce")

    avg_duration = pd.to_numeric(row.get("AvgDurationMinutes7D", 0), errors="coerce")
    max_duration = pd.to_numeric(row.get("MaxDurationMinutes7D", 0), errors="coerce")

    running_bonus_floor = 85 if is_running == 1 else 0
    is_queue_process = (trigger_type == "queue") or (has_queue_trigger == 1)

    if is_queue_process:
        if pd.notna(minutes_since_last_queue_txn):
            if minutes_since_last_queue_txn > 720:
                score -= 40
            elif minutes_since_last_queue_txn > 360:
                score -= 15
            elif minutes_since_last_queue_txn > 180:
                score -= 5

        if pd.notna(pending_items):
            if pending_items > 500:
                score -= 25
            elif pending_items > 200:
                score -= 15
            elif pending_items > 100:
                score -= 8
            elif pending_items > 20:
                score -= 3

        if failure_count_7d >= 10:
            score -= 15
        elif failure_count_7d >= 7:
            score -= 10
        elif failure_count_7d >= 5:
            score -= 5

        if pd.notna(days_since_last_activity):
            if days_since_last_activity >= 90:
                score -= 40
            elif days_since_last_activity >= 60:
                score -= 25
            elif days_since_last_activity >= 30:
                score -= 10

        score = max(score, 0)
        if is_running == 1:
            score = max(score, running_bonus_floor)
        return score

    if pd.notna(trigger_enabled) and int(trigger_enabled) == 0:
        return 10

    if is_running != 1 and pd.notna(expected):
        duration_buffer = max(avg_duration if pd.notna(avg_duration) else 0, 15)
        major_buffer = max(max_duration if pd.notna(max_duration) else duration_buffer, duration_buffer)

        warning_threshold = expected + duration_buffer + 100
        down_threshold = expected + major_buffer + 280

        if pd.notna(minutes_since_last_run):
            if minutes_since_last_run > down_threshold:
                score -= 30
            elif minutes_since_last_run > warning_threshold:
                score -= 10

        if pd.notna(minutes_since_last_success):
            if minutes_since_last_success > down_threshold:
                score -= 25
            elif minutes_since_last_success > warning_threshold:
                score -= 10

    if failure_count_7d >= 10:
        score -= 25
    elif failure_count_7d >= 7:
        score -= 15
    elif failure_count_7d >= 5:
        score -= 5

    if pd.notna(days_since_last_activity):
        if days_since_last_activity >= 90:
            score -= 40
        elif days_since_last_activity >= 60:
            score -= 25
        elif days_since_last_activity >= 30:
            score -= 10

    score = max(score, 0)
    if is_running == 1:
        score = max(score, running_bonus_floor)

    return score


def explain_state(row):
    process_state = row.get("ProcessState", "")
    trigger_type = normalize_text(row.get("TriggerTypeLatest")).lower()
    has_queue_trigger = pd.to_numeric(row.get("HasQueueTrigger", 0), errors="coerce")
    is_queue_process = (trigger_type == "queue") or (has_queue_trigger == 1)
    is_running = pd.to_numeric(row.get("IsCurrentlyRunning", 0), errors="coerce")
    minutes_since_last_queue_txn = row.get("MinutesSinceLastQueueTransaction")
    pending_items = pd.to_numeric(row.get("PendingItems", 0), errors="coerce")
    minutes_since_last_run = row.get("MinutesSinceLastRun")
    expected = pd.to_numeric(row.get("ExpectedIntervalMinutes"), errors="coerce")

    if process_state == "Running":
        return "Process is currently running"

    if process_state == "Stopped":
        if is_queue_process and pd.notna(minutes_since_last_queue_txn):
            return "Queue process appears stopped based on missing runtime movement"
        return "Process is stopped because trigger is disabled or no recent activity"

    if process_state == "Healthy":
        if is_queue_process:
            return f"Queue is operating normally with pending items={int(pending_items)}"
        return "Process is operating normally"

    if process_state == "Warning":
        if is_queue_process:
            return f"Queue activity delay or backlog detected (pending={int(pending_items)})"
        return f"Last run is delayed versus expected interval ({expected} min)"

    if process_state == "At Risk":
        if is_queue_process:
            return f"Queue is at risk due to old activity/backlog (pending={int(pending_items)})"
        return f"Process is significantly delayed compared to expected interval ({expected} min)"

    if process_state == "No Job History":
        return "No execution history found"

    if process_state == "LongTime Not Used":
        return "Process has no meaningful recent activity for a long time"

    if process_state == "Unknown":
        return "Expected schedule is not defined"

    return "No explanation available"


def explain_health_score(row):
    reasons = []
    trigger_type = normalize_text(row.get("TriggerTypeLatest")).lower()
    has_queue_trigger = pd.to_numeric(row.get("HasQueueTrigger", 0), errors="coerce")
    is_queue_process = (has_queue_trigger == 1) or (trigger_type == "queue")

    score = row.get("HealthScore", None)
    pending_items = pd.to_numeric(row.get("PendingItems", 0), errors="coerce")
    minutes_since_last_queue_txn = row.get("MinutesSinceLastQueueTransaction")
    failure_count_7d = pd.to_numeric(row.get("FailureCount7D", 0), errors="coerce")
    trigger_enabled = pd.to_numeric(row.get("TriggerEnabledLatest", 1), errors="coerce")

    if is_queue_process:
        if pd.notna(minutes_since_last_queue_txn):
            if minutes_since_last_queue_txn > 720:
                reasons.append(f"no queue transaction for {int(minutes_since_last_queue_txn)} minutes")
            elif minutes_since_last_queue_txn > 360:
                reasons.append(f"delayed queue activity ({int(minutes_since_last_queue_txn)} minutes)")

        if pending_items > 500:
            reasons.append(f"high queue backlog ({int(pending_items)} items)")
        elif pending_items > 200:
            reasons.append(f"moderate queue backlog ({int(pending_items)} items)")
        elif pending_items > 20:
            reasons.append(f"queue backlog exists ({int(pending_items)} items)")
    else:
        if pd.notna(trigger_enabled) and int(trigger_enabled) == 0:
            reasons.append("trigger is disabled")

    if failure_count_7d >= 10:
        reasons.append(f"high failures in last 7d ({int(failure_count_7d)})")
    elif failure_count_7d >= 7:
        reasons.append(f"elevated failures in last 7d ({int(failure_count_7d)})")

    if not reasons:
        return "No score penalties applied"

    return " | ".join(reasons)


# =========================================================
# BUILD PROCESS HEALTH
# =========================================================
def build_process_health(df_master, df_jobs_summary, df_queue_summary, logger):
    now_naive = pd.Timestamp.now(tz=LOCAL_TIMEZONE).tz_localize(None)

    queue_latest = df_queue_summary.copy()
    queue_latest["MinutesSinceLastQueueTransaction"] = (
        now_naive - queue_latest["LastQueueTransactionTime"]
    ).dt.total_seconds() / 60

    final_df = df_master.merge(
        queue_latest[
            [
                "QueueDefinitionId",
                "QueueName",
                "PendingItems",
                "FailedQueueItems",
                "SuccessfulQueueItems",
                "InProgressItems",
                "AbandonedItems",
                "RetriedItems",
                "LastQueueTransactionTime",
                "MinutesSinceLastQueueTransaction"
            ]
        ],
        on="QueueDefinitionId",
        how="left",
        suffixes=("", "_queue")
    )

    final_df = final_df.merge(df_jobs_summary, on="ReleaseId", how="left", suffixes=("", "_job"))

    if "QueueName" not in final_df.columns:
        final_df["QueueName"] = ""
    final_df["QueueName"] = final_df["QueueName"].fillna("")

    final_df["HasJobHistory"] = final_df["LastRunTime"].notna().astype(int)
    final_df["HasQueueTrigger"] = pd.to_numeric(final_df.get("HasQueueTrigger", 0), errors="coerce").fillna(0).astype(int)

    final_df["LastActivityTime"] = final_df[["LastRunTime", "LastQueueTransactionTime"]].max(axis=1)
    final_df["DaysSinceLastActivity"] = (now_naive - final_df["LastActivityTime"]).dt.days
    final_df["MinutesSinceLastRun"] = (now_naive - final_df["LastRunTime"]).dt.total_seconds() / 60
    final_df["MinutesSinceLastSuccess"] = (now_naive - final_df["LastSuccessTime"]).dt.total_seconds() / 60

    numeric_cols = [
        "PendingItems", "FailedQueueItems", "SuccessfulQueueItems", "InProgressItems",
        "AbandonedItems", "RetriedItems", "RunCount7D", "SuccessCount7D", "FailureCount7D",
        "TriggerEnabledLatest", "ItemsActivationThreshold", "ItemsPerJobActivationTarget",
        "MaxJobsForActivation", "AvgDurationMinutes7D", "MaxDurationMinutes7D", "IsCurrentlyRunning",
        "ExpectedIntervalMinutes", "ExpectedRunsPerDay", "ExpectedIntervalHours"
    ]
    for col in numeric_cols:
        if col not in final_df.columns:
            final_df[col] = 0
        final_df[col] = pd.to_numeric(final_df[col], errors="coerce")

    final_df["UsageStatus"] = final_df["DaysSinceLastActivity"].apply(classify_usage)
    final_df["ProcessState"] = final_df.apply(classify_process_state, axis=1)
    final_df["HealthScore"] = final_df.apply(calculate_health_score, axis=1)
    final_df["Health_Explanation"] = final_df.apply(explain_health_score, axis=1)
    final_df["State_Explanation"] = final_df.apply(explain_state, axis=1)

    sort_map = {
        "At Risk": 1,
        "Warning": 2,
        "Running": 3,
        "Healthy": 4,
        "Stopped": 5,
        "LongTime Not Used": 6,
        "Disabled": 7,
        "No Job History": 8,
        "Unknown": 9
    }
    final_df["StateSortOrder"] = final_df["ProcessState"].map(sort_map).fillna(99)
    
    final_df = final_df.sort_values(["StateSortOrder", "HealthScore", "ProcessName"], ascending=[True, True, True])
# Rename Severity to Criticality for health output
    final_df["IsEnabled"] = final_df["TriggerEnabledLatest"]
    if "Severity" in final_df.columns:
        final_df = final_df.rename(columns={"Severity": "Criticality"})
    logger.info(f"Process health built | rows={len(final_df)}")
    return final_df


# =========================================================
# SAVE OUTPUTS
# =========================================================
def save_outputs(base, config, df_master, df_jobs, df_queue_summary, df_health, logger):
    out_dir = os.path.join(base, "outputs", "process_health")
    shared_dir = config.get("SharedOutputDir", "./outputs/shared")

    workbook_path = config.get("MasterWorkbookPath", "./outputs/Alert_Process_Severity_Mapping.xlsx")

    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(shared_dir, exist_ok=True)

    files = {
        "Process_Master_config.csv": df_master,
        "Queues_Summary.csv": df_queue_summary,
        "UiPath_Process_Health.csv": df_health
    }

    for filename, df in files.items():
        local_path = os.path.join(out_dir, filename)
        shared_path = os.path.join(shared_dir, filename)

        export_df = df.copy()
        export_df.columns = export_df.columns.str.strip()

        export_df.to_csv(local_path, index=False, encoding="utf-8-sig")
        export_df.to_csv(shared_path, index=False, encoding="utf-8-sig")

        logger.info(f"Exported: {local_path} | rows={len(export_df)}")
        logger.info(f"Exported: {shared_path} | rows={len(export_df)}")

    # update only UIPATH sheet inside existing workbook
    save_uipath_sheet_only(workbook_path, df_master, logger, sheet_name="UIPATH")



# =========================================================
# MAIN
# =========================================================

# -----------------------------------------------------------------------------
# PROCESS HEALTH ETL FLOW
# -----------------------------------------------------------------------------
# 1. Load configuration and initialize logging
# 2. Extract releases (full inventory)
# 3. Load business mapping rules
# 4. Load jobs history from curated data lake
# 5. Extract queue and schedule metadata
# 6. Build unified master process mapping
# 7. Aggregate job execution metrics
# 8. Calculate process health and state
# 9. Export outputs for dashboards and monitoring
# -----------------------------------------------------------------------------

def run_process_health_etl():
    config_path = os.getenv("RPA_HEALTH_CONFIG_PATH", "./config/config.xlsx")
    business_file = os.getenv("RPA_BUSINESS_MAPPING_PATH", "./data/business_mapping.xlsx")



    config = load_settings(config_path, 'Settings', 'Name', 'Value')
    required_keys = [
    "SQL_Server",
    "SQL_Database",
    "SQL_UID",
    "SQL_PWD",
    "BASE_DATA_LAKE"
    ]

    missing = [k for k in required_keys if k not in config or str(config[k]).strip() == ""]
    if missing:
        raise ValueError(f"Missing required config keys: {missing}")
    
    server = config["SQL_Server"]
    database = config["SQL_Database"]
    username = config["SQL_UID"]
    password = config["SQL_PWD"]

    output_dir = config.get("OutputDir", "./outputs")
    shared_dir = config.get("SharedOutputDir", "./outputs/shared")
    base = config["BASE_DATA_LAKE"]
    curated_db = os.path.join(base, "curated", "rpa_lake.db")

    log_file_path = os.path.join(
    config.get("LogDir", "./logs"),
    "ETL_UiPath_Process_Health",
    f"ETL_UiPath_Process_Health_{datetime.now().strftime('%m-%d-%Y')}.log"
)

    logger_main, logger_init, logger_process, fileHandler = setup_loggers(log_file_path)

    sql_conn = None
    try:
        logger_main.info("=======================================")
        logger_main.info("ETL UiPath Unified Master started")
        logger_main.info("=======================================")



        logger_init.info("Init Settings Success")

        conn_str = build_sql_conn_str(server, database, username, password)
        sql_conn = pyodbc.connect(conn_str)

        # 1) full inventory from releases
        df_releases = extract_releases(sql_conn, logger_process)

        # 2) business mapping
        df_business = load_business_mapping(business_file, logger_process)

        # 3) jobs from curated lake
        df_jobs = load_jobs_from_curated(curated_db, logger_process, lookback_days=90)

        # 4) technical enrichments
        df_queue_summary = extract_queue_summary(sql_conn, logger_process)
        df_schedules = extract_process_schedule_map(sql_conn, logger_process)

        df_schedules = enrich_schedule_fields(df_schedules, logger_process)
        df_trigger_map = build_process_trigger_map(df_schedules, df_releases, logger_process)

        # 5) build master from ALL releases
        df_master = build_master_process_mapping(
            df_releases=df_releases,df_business=df_business,
            df_trigger_map=df_trigger_map,logger=logger_process)

        # 6) jobs summary
        df_jobs_summary = build_jobs_summary(df_jobs, logger_process)

        # 7) health from full master
        df_health = build_process_health(df_master, df_jobs_summary, df_queue_summary, logger_process)
        logger_main.info(f"Master rows: {len(df_master)}")
        logger_main.info(f"Health rows: {len(df_health)}")

        # 8) save outputs
        save_outputs(base, config, df_master, df_jobs, df_queue_summary, df_health, logger_main)

        logger_main.info(f"Total releases in inventory: {len(df_master)}")
        logger_main.info(f"Processes with business mapping: {int(df_master['InBusinessMapping'].sum())}")
        logger_main.info(f"Processes missing business mapping: {int(df_master['MissingBusinessMapping'].sum())}")

        logger_main.info("ETL UiPath Unified Master finished successfully ✅")

    except Exception as e:
        logger_main.exception(f"ETL UiPath Unified Master failed: {e}")
        raise

    finally:
        if sql_conn is not None:
            try:
                sql_conn.close()
            except Exception:
                pass

        if fileHandler:
            fileHandler.close()
        logging.shutdown()


if __name__ == "__main__":
    run_process_health_etl()